import os
import re
import json
import pandas as pd
from openpyxl import load_workbook

# Read firm marketing plan
path = "./input-ready/"
files = os.listdir(path)
files = [file for file in files if file.endswith('.xlsx')]
file = files[0]

# Function to create the filename and write the initial empty lines
def make_filename(file):
  # Load an Excel file
  workbook = load_workbook(path + file, data_only=True)
  sheet = workbook["Marketing Plan"]
  
  # Read firm, market area, season data
  firm_number = str(re.search(r'\d', file).group())
  market_area = sheet["E44"].value
  restaurant_name = sheet["L2"].value
  season = sheet["L4"].value
  
  # Create needed variables
  f_name = "RFI" + firm_number + market_area + ".TXT"  # Concatenate the parts to form the file name
  
  # Open the file in write mode and add three empty lines
  with open(f_name, "w") as file:
      file.write("\r\n\r\n\r\n")
  
  print(f"File '{f_name}' has been created with three empty lines.")
  
  # Create the line containing the restaurant name with appropriate spacing
  restaurant_string = restaurant_name.center(42)
  restaurant_string = restaurant_string.ljust(84) + "LS        B1"

  # Append the restaurant name to the file
  with open(f_name, "a") as file:  # Open the file in append mode
      file.write(restaurant_string + "\r\n")  # Add the string with a newline
  
  print(f"'{restaurant_string}'")  # Prints the result with quotes to visualize spaces
  
  return f_name

# Function to create the menu decisions lines
def make_menu(file, f_name):
  # Load an Excel file
  workbook = load_workbook(path + file, data_only=True)
  sheet = workbook["Marketing Plan"]
  
  # Read menu data
  menu_rows = [10 + 2 * i for i in range(8)]
  menu_items = [sheet[f"A{row}"].value for row in menu_rows]
  menu_preps = [sheet[f"E{row}"].value for row in menu_rows]
  pnps = [sheet[f"C{row}"].value for row in menu_rows]
  portions = [sheet[f"G{row}"].value for row in menu_rows]
  mkt = [sheet[f"J{row}"].value for row in menu_rows]
  prices = [sheet[f"L{row}"].value for row in menu_rows]
  forecast = [0, 0, 0, 0, 0, 0, 0, 0]

  # Menu saved as a pandas dataframe
  menu = pd.DataFrame({
      "menu_item": menu_items,
      "menu_prep": menu_preps,
      "pnp": pnps,
      "portion": portions,
      "marketing": mkt,
      "price": prices,
      "food_cost": [0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000],
      "forecast": forecast,
  })
  menu = menu[menu["menu_item"].notna()]
   
  # Enforce specific data types
  menu = menu.astype({
      "menu_item": "string",      # Enforce as string type
      "menu_prep": "string",      # Enforce as string type
      "pnp": "int",             # Nullable integer type
      "portion": "float",         # Float type
      "marketing": "int",       # Integer type
      "price": "float",           # Float type
      "food_cost": "float",       # Float type
      "forecast": "int",       # Integer type
  })
  
  # Append menu decisions
  # Create each line and append it
  menu_item_line = []
  menu_item_line_count = 0
  for _, row in menu.iterrows():
      # Format the string with exact positions
      menu_item_line = (
          f"{row['menu_item']:<32}"        # menu_item starts at the beginning
          f"{row['menu_prep']:<20}"        # menu_prep starts at character 32
          f"{row['pnp']:>1}"               # pnp starts at character 52
          f"{row['portion']:>10.1f}"       # portion starts at character 58
          f"{row['marketing']:>10}"        # marketing starts at character 68
          f"{row['price']:>10.2f}"         # price starts at character 79 ends at 83
          f"{row['food_cost']:>10.3f}"     # food cost ends at character 93
          f"{row['forecast']:>10}"          # forecast is right aligned ending at 103
      )
      with open(f_name, "a") as file:  # Open the file in append mode
        file.write(menu_item_line + "\r\n")  # Add the string with a newline
        menu_item_line_count += 1  # Increment the line count
  
  # Add the required number of empty lines to get to 10 menu items
  with open(f_name, "a") as file:
      for _ in range(10 - menu_item_line_count):  # Loop for the required number of lines
          padded_line = f"{'':<88}0.000\r\n"  # Pad with spaces to 87 characters, then add "0.000"
          file.write(padded_line)
          
  print(f"File '{f_name}' now contains 10 lines, with {menu_item_line_count} menu items and the remaining as blank lines.")
  
  return menu

# Function to create the operations decisions lines
def make_ops(file, f_name):
  # Load an Excel file
  workbook = load_workbook(path + file, data_only=True)
  sheet = workbook["Marketing Plan"]
  if "Financial Plan" in workbook.sheetnames:
    sheet2 = workbook["Financial Plan"]
  
  # Read firm, market area, season data
  fte = sheet["E27"].value
  training = sheet["E29"].value  
  big4 = sheet["E33"].value
  music = sheet["E35"].value
  maintenance = sheet["E37"].value
  infoguest = 1 if sheet["E42"].value == "Yes" else 0
  infofin = 1 if sheet["G42"].value == "Yes" else 0
  infoprod = 1 if sheet["J42"].value == "Yes" else 0
  
  # Create operations decisions line
  ops_decisions = (
    f"{'Operations':<43}"   # Operations left aligned
    f"{fte:>10.1f}"         # full time equivalents are right aligned to position 53
    f"{training:>10.2f}"    # training is 10 characters right aligned 2 decimal
    f"{big4:>10.2f}"         # big4 is 10 characters right aligned 2 decimal
    f"{music:>10}"          # music is 10 characters right aligned integer
    f"{maintenance:>10}" # maintenance is 10 characters right aligned integer
    f"{infoguest:>10}"           # guest satisfaction 10 characters right aligned
    f"{infofin:>10}"           # financial report 10 characters right aligned
    f"{infoprod:>10}"           # product report 10 characters right aligned
    f"{'0':>10}"           # unknown padding 0 value is 10 characters right aligned
    f"{'12000':>10}"       # unknown padding 12000 value is 10 characters right aligned
)

  # Define the additional lines to be added
  additional_lines1 = [
      "Management                                        1.5       0.0       0.0         0         0         0         0         0         0         0",
      "Supervisory                                       1.0       0.0       0.0         0         0         0         0         0         0         0",
      "Service                                          10.0       0.0       0.0         0         0",
      "Support                                           9.0       0.0       0.0         0         0",
  ]
  
  fin_empty_lines = str("Finance                                             0         0         0         0         0         0         0         0         0         0")
  
  additional_lines2 = [      
      "Assets - Investments                                0         0",
      "       - Leases                                     0         0",
      "       - Borrowings                                 0         0",
      "Variances",
  ]
  
  # Write the additional lines to the file
  with open(f_name, "a") as f:
      for line in additional_lines1:
          f.write(line + "\r\n")
      f.write(ops_decisions + "\r\n")
      try:
         print(sheet2)
         fin_line = make_fin(file)
         f.write(fin_line + "\r\n")
      except NameError:
         f.write(fin_empty_lines + "\r\n")
      for line in additional_lines2:
          f.write(line + "\r\n")

# Function to create the financial decisions line
def make_fin(file):
  # Load an Excel file
  workbook = load_workbook(path + file, data_only=True)
  sheet = workbook["Financial Plan"]
  
  # Read firm, market area, season data
  timedep = str(sheet["D7"].value) + str(sheet["E7"].value)
  certdep = str(sheet["D9"].value) + str(sheet["E9"].value)
  notepay = str(sheet["D11"].value) + str(sheet["E11"].value)
  capital = str(sheet["D13"].value) + str(sheet["E13"].value)
  dividend = str(sheet["D15"].value) + str(sheet["E15"].value)
  
  # Create financial decisions line
  fin_decisions = (
    f"{'Finance':<43}"   # Financial left aligned, padded to 43 characters
    f"{timedep:>10}"         # sales are right aligned to position 53
    f"{certdep:>10}"    # food cost is 10 characters right aligned 2 decimal
    f"{notepay:>10}"         # beverage cost is 10 characters right aligned 2 decimal
    f"{capital:>10}"          # labor cost is 10 characters right aligned integer
    f"{dividend:>10}" # other costs is 10 characters right aligned integer
    f"{'0':>10}"           # unknown padding 0 value is 10 characters right aligned
    f"{'0':>10}"           # unknown padding 0 value is 10 characters right aligned
    f"{'0':>10}"           # unknown padding 0 value is 10 characters right aligned
    )
  return fin_decisions

# Main execution
for file in files:
  print(file)
  f_name = make_filename(file)
  menu = make_menu(file, f_name)
  print(menu)
  make_ops(file, f_name)
  
  # Close file with EOF character
  with open(f_name, "a") as file:
    file.write("\r\n")

